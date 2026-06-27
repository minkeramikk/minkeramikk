#!/usr/bin/env python3
"""
i18n-export.py — esporta tutte le traduzioni (NO/EN) in un foglio Excel
per la revisione da parte del cliente (Alessio).

Legge:   web/src/i18n/messages/en.json + no.json
Produce: un .xlsx con una riga per chiave (key | namespace | EN | NO | placeholder | note),
         colonne Key/Namespace/Placeholder bloccate, EN/NO/Note editabili.

Round-trip: il cliente edita SOLO le colonne English/Norsk/Note; poi `i18n-import.py`
(da fare) rilegge il file, valida i placeholder e riscrive i JSON.

Uso:
    python3 scripts/i18n-export.py [output.xlsx]
Default output: ../docs/client/minkeramikk-traduzioni.xlsx
"""
from __future__ import annotations
import json
import re
import sys
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side, Protection
from openpyxl.utils import get_column_letter

# --- percorsi (relativi alla posizione dello script) ---
SCRIPT_DIR = Path(__file__).resolve().parent          # web/scripts
WEB_DIR = SCRIPT_DIR.parent                            # web
REPO_ROOT = WEB_DIR.parent                             # repo root
MESSAGES_DIR = WEB_DIR / "src" / "i18n" / "messages"
DEFAULT_OUT = REPO_ROOT / "docs" / "client" / "minkeramikk-traduzioni.xlsx"

ICU_RE = re.compile(r"\{[^}]+\}")


def flatten(d: dict, prefix: str = "") -> dict[str, str]:
    """Appiattisce un dict annidato in chiavi dot-notation -> valore stringa."""
    out: dict[str, str] = {}
    for k, v in d.items():
        key = f"{prefix}{k}"
        if isinstance(v, dict):
            out.update(flatten(v, key + "."))
        else:
            out[key] = "" if v is None else str(v)
    return out


def placeholders(text: str) -> str:
    found = ICU_RE.findall(text or "")
    # ordine stabile, senza duplicati
    seen: list[str] = []
    for f in found:
        if f not in seen:
            seen.append(f)
    return " ".join(seen)


def main() -> int:
    out_path = Path(sys.argv[1]).resolve() if len(sys.argv) > 1 else DEFAULT_OUT

    en_path = MESSAGES_DIR / "en.json"
    no_path = MESSAGES_DIR / "no.json"
    if not en_path.exists() or not no_path.exists():
        print(f"ERRORE: dizionari non trovati in {MESSAGES_DIR}", file=sys.stderr)
        return 1

    en = flatten(json.loads(en_path.read_text(encoding="utf-8")))
    no = flatten(json.loads(no_path.read_text(encoding="utf-8")))

    # unione delle chiavi: cosi' eventuali chiavi mancanti in una lingua restano visibili
    all_keys = sorted(set(en) | set(no), key=lambda k: (k.split(".")[0], k))

    # --- workbook ---
    wb = Workbook()

    # Foglio istruzioni
    ws_info = wb.active
    ws_info.title = "Istruzioni"
    instr = [
        ("Come revisionare le traduzioni", True),
        ("", False),
        ("1. Vai al foglio \"Traduzioni\".", False),
        ("2. Modifica SOLO le colonne \"English (EN)\", \"Norsk (NO)\" e \"Note\".", False),
        ("3. NON toccare le colonne \"Key\" e \"Namespace\": servono al sistema (sono bloccate).", False),
        ("4. I segnaposto tra graffe come {code}, {count}, {step} vanno LASCIATI identici", False),
        ("   nel testo: sono valori che il sito inserisce in automatico. Puoi spostarli ma non", False),
        ("   rinominarli ne' rimuoverli. La colonna \"Placeholder\" ricorda quali ci sono.", False),
        ("5. Una cella vuota in NO o EN = traduzione mancante: va compilata.", False),
        ("6. Quando hai finito, rimanda il file: lo reimportiamo noi nel sito.", False),
        ("", False),
        ("Nota: il foglio Traduzioni e' protetto per evitare modifiche accidentali alle chiavi.", False),
    ]
    for i, (text, bold) in enumerate(instr, start=1):
        c = ws_info.cell(row=i, column=1, value=text)
        if bold:
            c.font = Font(bold=True, size=14)
    ws_info.column_dimensions["A"].width = 95

    # Foglio traduzioni
    ws = wb.create_sheet("Traduzioni")
    headers = ["Key", "Namespace", "English (EN)", "Norsk (NO)", "Placeholder (non rimuovere)", "Note"]
    header_fill = PatternFill("solid", fgColor="2F4858")
    header_font = Font(bold=True, color="FFFFFF")
    thin = Side(style="thin", color="D0D0D0")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    locked_fill = PatternFill("solid", fgColor="F2F2F2")

    for col, h in enumerate(headers, start=1):
        c = ws.cell(row=1, column=1 if col == 1 else col, value=h)
        c.fill = header_fill
        c.font = header_font
        c.alignment = Alignment(vertical="center", wrap_text=True)
        c.border = border

    # colonne EDITABILI (1-based): English=3, Norsk=4, Note=6
    editable_cols = {3, 4, 6}

    for r, key in enumerate(all_keys, start=2):
        ns = key.split(".")[0]
        en_val = en.get(key, "")
        no_val = no.get(key, "")
        ph = placeholders(en_val) or placeholders(no_val)
        row_vals = [key, ns, en_val, no_val, ph, ""]
        for col, val in enumerate(row_vals, start=1):
            cell = ws.cell(row=r, column=col, value=val)
            cell.border = border
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            if col in editable_cols:
                cell.protection = Protection(locked=False)
            else:
                cell.protection = Protection(locked=True)
                cell.fill = locked_fill

    # larghezze e formattazione
    widths = {1: 34, 2: 16, 3: 50, 4: 50, 5: 22, 6: 30}
    for col, w in widths.items():
        ws.column_dimensions[get_column_letter(col)].width = w
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{len(all_keys) + 1}"

    # protezione foglio: tutto bloccato tranne le celle sbloccate sopra
    ws.protection.sheet = True
    ws.protection.enable()
    ws.protection.formatCells = False

    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_path)

    missing_no = sum(1 for k in all_keys if not no.get(k))
    missing_en = sum(1 for k in all_keys if not en.get(k))
    print(f"OK  {len(all_keys)} chiavi esportate -> {out_path}")
    print(f"    mancanti EN: {missing_en} · mancanti NO: {missing_no}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
